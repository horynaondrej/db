#!/usr/bin/env python3
"""Modul pro výběr dat z SQLite databáze.

Obsahuje třídu a metody pro připojení k databázi, provádění SQL dotazů,
zobrazení výsledků a export dat do souborů.
"""

import logging
import os
import shutil
import importlib.util
import duckdb
import pandas
from openpyxl import load_workbook

from master_config import SLOZKA

# vytvoř absolutní cestu k modulu config.py
cesta_k_modulu = os.path.join(os.path.dirname(__file__), SLOZKA, "config.py")

# vytvoř import spec
spec = importlib.util.spec_from_file_location("config", cesta_k_modulu)
config = importlib.util.module_from_spec(spec)
spec.loader.exec_module(config)


class VyberciDat:
    """Třída pro práci s databází"""

    def __init__(self):
        self.cesta = os.path.dirname(__file__)  # Relativní cesta k souborům
        self.db = os.path.join(self.cesta, config.DB)  # Cesta k databázi
        self.con = None  # Připojení k databázi
        self.cur = None  # Kurzory pro SQL dotazy
        self.sablona = os.path.join(self.cesta, "sablona.xlsx")
        self.vystup = os.path.join(self.cesta, SLOZKA, "result.xlsx")  # Výstupní soubor
        self.vysledne = pandas.DataFrame()  # Uchování výsledků dotazů

        # Nastavení knihovny pandas pro zobrazení dat
        pandas.set_option("display.max_columns", 3000)
        pandas.set_option("display.max_rows", 500)
        pandas.set_option("display.expand_frame_repr", False)
        pandas.set_option("max_colwidth", 3000)
        pandas.set_option("display.width", 3000)
        pandas.options.display.float_format = "{:20,.2f}".format

    def pripoj_se_k_databazi(self):
        """připojí se k databázi"""
        if os.path.isfile(self.db):
            self.con = duckdb.connect(self.db)
        else:
            logging.info("Databáze nenalezena")

    def vyber_data_z_databaze(self, dotaz):
        """provede SQL dotaz na databázi"""
        if self.con:
            return self.con.execute(dotaz).fetchdf()
        logging.error("Nelze provést dotaz, připojení k databázi nebylo úspěšné.")
        return None

    def odpoj_se_od_databaze(self):
        """potvrdí příkazy a ukončí připojení k databázi"""
        if self.con:
            self.con.commit()
            self.con.close()

    def zkopiruj_xlsx_sablonu(self):
        """zkopíruje šablonu excelu, do které se data uloží
        aby se zachovalo nastavené formátování
        """
        shutil.copy2(self.sablona, self.vystup)

    def uloz_data_do_xlsx(self, df):
        """ulož dataframe do xlsx"""
        res = [df.columns.tolist()] + df.values.tolist()

        # Načtení existujícího Excelového souboru
        book = load_workbook(self.vystup)
        sheet = book["List1"]

        logging.info("Exportuji do excelu..")

        for row_index, row in enumerate(res, start=1):
            for col_index, value in enumerate(row, start=1):
                sheet.cell(row=row_index, column=col_index, value=value)

        # Uložení změn
        book.save(self.vystup)


def main():
    """Hlavní funkce skriptu, která inicializuje a spouští procesy."""
    fmt = "%(asctime)s: %(message)s"
    logging.basicConfig(format=fmt, level=logging.INFO, datefmt="%H:%M:%S")

    logging.info("Spuštění skriptu")

    # Pole pro uchování jednotlivých dotazů
    dotazy = []
    _dotazy = []

    # Dotaz na data z tabulky track
    _dotazy.append(
        """
        select datum_ucetni_rokmesic, sum(naklady_opravy_a_udrzba) as mer 
        from pm
        group by datum_ucetni_rokmesic

    """
    )

    dotazy.append(
        """
        select * 
        from t

    """
    )

    # Projde seznam dotazů a provede je postupně
    for i in dotazy:
        vyberci_dat = VyberciDat()
        vyberci_dat.pripoj_se_k_databazi()
        df = vyberci_dat.vyber_data_z_databaze(i)
        print("\n", df, "\n")
        vyberci_dat.zkopiruj_xlsx_sablonu()
        vyberci_dat.uloz_data_do_xlsx(df)
        vyberci_dat.odpoj_se_od_databaze()

    logging.info("Ukončení skriptu")


# Hlavní vlákno skriptu
if __name__ == "__main__":
    main()
